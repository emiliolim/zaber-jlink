"""
parses through the log files and extracts
relevant time data for analysis.
To run: 
python logparse.py path/to/logfile.log --debug
--debug flag will print out negative deltas with their source pairs for debugging purposes.

to plot:
python logparse.py rtt-tests/Run1.log --plot
--excel plots all time stamps
--limit TIME limits the time values to a max TIME value
"""

from __future__ import annotations

import argparse
import re
import statistics
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except ImportError:
    plt = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

TIME_PATTERN = re.compile(r"TIME:\s*([-+]?\d+(?:\.\d+)?)\s*\(sec\)", re.IGNORECASE)
A_PATTERN = re.compile(r"\bA:\s*([\-\d,]+)")
C_PATTERN = re.compile(r"\bC:\s*([\-\d,]+)")


def parse_records_from_file(file_path: Path, time_limit: float = 180.0) -> List[dict]:
    """Parse records from the RTT log file.

    Each record is a dict: {"time": float, "A": Optional[List[int]], "C": List[List[int]]}
    A record starts at a line containing `TIME:` and may include following `A:` and `C:` lines.
    Parsing stops once a `TIME` value exceeds `time_limit`.
    """
    records: List[dict] = []
    current: Optional[dict] = None

    with file_path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            tmatch = TIME_PATTERN.search(line)
            if tmatch:
                # push previous record (avoid duplicate consecutive times)
                if current is not None:
                    if not records or records[-1].get("time") != current.get("time"):
                        records.append(current)

                try:
                    time_value = float(tmatch.group(1))
                except ValueError:
                    current = None
                    continue

                if time_value > time_limit:
                    break

                current = {"time": time_value, "A": None, "C": []}
                continue

            if current is None:
                # skip lines until we see a TIME
                continue

            amatch = A_PATTERN.search(line)
            if amatch:
                try:
                    vals = [int(x) for x in amatch.group(1).split(",") if x.strip()]
                except ValueError:
                    vals = []
                current["A"] = vals

            cmatch = C_PATTERN.search(line)
            if cmatch:
                try:
                    vals = [int(x) for x in cmatch.group(1).split(",") if x.strip()]
                except ValueError:
                    vals = []
                current["C"].append(vals)

    # append last record if present (avoid duplicate consecutive times)
    if current is not None:
        if not records or records[-1].get("time") != current.get("time"):
            records.append(current)

    return records


def compute_deltas(times: Iterable[float]) -> List[float]:
    time_list = list(times)
    return [time_list[i] - time_list[i - 1] for i in range(1, len(time_list))]


def find_negative_deltas(times: Iterable[float]) -> List[tuple[float, float, float]]:
    time_list = list(times)
    negative_pairs: List[tuple[float, float, float]] = []
    for previous, current in zip(time_list, time_list[1:]):
        delta = current - previous
        if delta < 0:
            negative_pairs.append((previous, current, delta))
    return negative_pairs


def find_negative_delta_records(records: Iterable[dict]) -> List[tuple[dict, dict, float]]:
    recs = list(records)
    negative_pairs: List[tuple[dict, dict, float]] = []
    for prev, curr in zip(recs, recs[1:]):
        delta = curr["time"] - prev["time"]
        if delta < 0:
            negative_pairs.append((prev, curr, delta))
    return negative_pairs


def compute_statistics(values: Iterable[float]) -> Tuple[Optional[float], Optional[float]]:
    values_list = list(values)
    if not values_list:
        return None, None
    if len(values_list) == 1:
        return values_list[0], 0.0
    return statistics.mean(values_list), statistics.pstdev(values_list)


def format_value(value: Optional[float]) -> str:
    return "n/a" if value is None else f"{value:.6f}"


def plot_deltas(times: List[float], deltas: List[float], title: str, output_path: Path, yzoom: Optional[float] = None) -> None:
    if plt is None:
        print("matplotlib is not installed; cannot create plot.")
        return

    x_values = times[1:]
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(x_values, deltas, marker="o", linestyle="-", color="tab:blue")
    ax.axhline(0.0, color="tab:red", linestyle="--", linewidth=0.8)
    ax.set_xlabel("TIME (sec)")
    ax.set_ylabel("Delta TIME (sec)")
    ax.set_title(title)
    # Make y-axis hyperspecific: center around mean delta with a tight span.
    mean_delta, deviation = compute_statistics(deltas)
    if mean_delta is None:
        mean_delta = 0.0
    if deviation is None:
        deviation = 0.0

    # If user provided an explicit yzoom, use it as half-range; otherwise use max(5*sigma, 0.002s)
    if yzoom is not None:
        span = float(yzoom)
    else:
        span = max(5 * (deviation or 0.0), 0.002)

    ax.set_ylim(mean_delta - span, mean_delta + span)
    try:
        import matplotlib.ticker as mticker
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.6f"))
        ax.yaxis.set_minor_locator(mticker.AutoMinorLocator())
    except Exception:
        pass
    ax.grid(True, linestyle="--", alpha=0.5)
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)
    print(f"Plot saved to {output_path}")


def export_to_excel(times: List[float], logfile_path: Path) -> None:
    """Export timestamps to an Excel file with one column."""
    if openpyxl is None:
        print("openpyxl is not installed; cannot create Excel file.")
        print("Install it with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timestamps"
    ws["A1"] = "TIME (sec)"

    for idx, time_value in enumerate(times, start=2):
        ws[f"A{idx}"] = time_value

    output_file = logfile_path.with_suffix("").with_name(f"{logfile_path.stem}-timestamps.xlsx")
    wb.save(output_file)
    print(f"Excel file saved to {output_file}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse TIME entries and compute timing statistics.")
    parser.add_argument("logfile", type=Path, help="Path to the RTT log file to parse.")
    parser.add_argument("--limit", type=float, default=180.0, help="Maximum TIME value to include in seconds.")
    parser.add_argument("--debug", action="store_true", help="Print negative TIME deltas with their source pairs.")
    parser.add_argument("--plot", action="store_true", help="Save a plot of delta over time to a PNG file.")
    parser.add_argument("--yzoom", type=float, default=None, help="Half-range for y-axis in seconds (e.g. 0.001). Overrides automatic scaling.")
    parser.add_argument("--excel", action="store_true", help="Export timestamps to an Excel file.")
    args = parser.parse_args()

    records = parse_records_from_file(args.logfile, time_limit=args.limit)
    if not records:
        print("No TIME entries found before the limit.")
        return

    times = [r["time"] for r in records]
    deltas = compute_deltas(times)
    mean_delta, deviation = compute_statistics(deltas)

    print(f"Parsed {len(times)} TIME entries up to {args.limit} seconds.")
    print(f"First TIME: {format_value(times[0])} sec")
    print(f"Last TIME: {format_value(times[-1])} sec")
    print(f"Mean difference (delta): {format_value(mean_delta)} sec")
    print(f"Deviation of deltas: {format_value(deviation)} sec")

    if deltas:
        print(f"Min delta: {format_value(min(deltas))} sec")
        print(f"Max delta: {format_value(max(deltas))} sec")

    if args.debug:
        negative_pairs = find_negative_delta_records(records)
        if not negative_pairs:
            print("No negative deltas found.")
        else:
            print("Negative TIME deltas found:")
            for prev_rec, curr_rec, delta in negative_pairs:
                prev_time = prev_rec.get("time")
                curr_time = curr_rec.get("time")
                print(f"  previous={format_value(prev_time)} sec, current={format_value(curr_time)} sec, delta={format_value(delta)} sec")
                if prev_rec.get("A") is not None or curr_rec.get("A") is not None:
                    print(f"    prev A={prev_rec.get('A')}, curr A={curr_rec.get('A')}")
                if prev_rec.get("C") or curr_rec.get("C"):
                    print(f"    prev C={prev_rec.get('C')}, curr C={curr_rec.get('C')}")

    if args.plot:
        output_file = args.logfile.with_suffix("").with_name(f"{args.logfile.stem}-delta.png")
        plot_deltas(times, deltas, f"Delta over time for {args.logfile.name}", output_file, yzoom=args.yzoom)

    if args.excel:
        export_to_excel(times, args.logfile)


if __name__ == "__main__":
    main()
