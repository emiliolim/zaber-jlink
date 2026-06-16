"""
Parse Excel files and analyze the TIME column.
python pyparse.py --debug --limit 180
python pyparse.py --plot-cap
"""

from __future__ import annotations

import argparse
import re
import statistics
from pathlib import Path
from typing import Iterable, Iterator, List, Optional, Tuple

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
except ImportError:
    plt = None

import openpyxl
import json

TIME_VALUE_PATTERN = re.compile(r"([-+]?\d+(?:\.\d+)?)")


def parse_time_value(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None

    match = TIME_VALUE_PATTERN.search(text)
    if not match:
        return None

    try:
        return float(match.group(1))
    except ValueError:
        return None


def find_time_column(header_row) -> Optional[int]:
    for index, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().upper()
        # accept 'TIME', 'TIME(s)', or any header that starts with 'TIME'
        if name.startswith("TIME"):
            return index
    return None


def find_cap_columns(header_row) -> dict[str, int]:
    cap_columns: dict[str, int] = {}
    for index, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip().upper()
        # Accept headers like 'CAP1', 'CAP1(PF)', or 'CAP 1 (pF)'
        m = re.match(r"^CAP\s*(\d+)", name)
        if not m:
            continue
        cap_index = int(m.group(1))
        if 1 <= cap_index <= 8:
            cap_columns[f"CAP{cap_index}"] = index
    return cap_columns


def parse_sheet_data_from_xlsx(file_path: Path, time_limit: Optional[float] = None) -> tuple[list[float], dict[str, list[Optional[float]]]]:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    header_row_index = None
    time_col_index = None
    cap_col_indices: dict[str, int] = {}

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        time_col_index = find_time_column(row)
        if time_col_index is not None:
            cap_col_indices = find_cap_columns(row)
            header_row_index = row_index
            break

    if time_col_index is None:
        raise ValueError(f"No TIME column found in {file_path.name}")

    cap_data: dict[str, list[Optional[float]]] = {f"CAP{i}": [] for i in range(1, 9)}
    times: list[float] = []

    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        cell_value = row[time_col_index] if time_col_index < len(row) else None
        time_value = parse_time_value(cell_value)
        if time_value is None:
            continue

        if time_limit is not None and time_value > time_limit:
            break

        times.append(time_value)
        for channel, column_index in cap_col_indices.items():
            cell_value = row[column_index] if column_index < len(row) else None
            cap_data[channel].append(parse_time_value(cell_value))

        for channel in cap_data:
            if channel not in cap_col_indices:
                cap_data[channel].append(None)

    return times, cap_data


def get_xlsx_files(folder: Path) -> Iterator[Path]:
    yield from sorted(folder.glob("*.xlsx"))


def compute_deltas(times: Iterable[float]) -> List[float]:
    time_list = list(times)
    return [time_list[i] - time_list[i - 1] for i in range(1, len(time_list))]


def find_negative_deltas(times: Iterable[float]) -> List[tuple[float, float, float]]:
    time_list = list(times)
    negative_pairs: List[tuple[float, float, float]] = []
    for previous, current in zip(time_list, time_list[1:]):
        delta = current - previous
        if delta < 0:
            negative_pairs.append((previous, current, delta))
    return negative_pairs


def compute_statistics(values: Iterable[float]) -> Tuple[Optional[float], Optional[float]]:
    values_list = list(values)
    if not values_list:
        return None, None
    if len(values_list) == 1:
        return values_list[0], 0.0
    return statistics.mean(values_list), statistics.pstdev(values_list)


def compute_mean_peak_trough(values: Iterable[Optional[float]]) -> Optional[float]:
    """Compute the mean peak-to-trough amplitude from a time series.

    Method: find local extrema (strict or equal neighbor allowed) and compute
    absolute differences between consecutive extrema; return their mean.
    Returns None if not enough extrema are found.
    """
    vals = [v for v in values if v is not None]
    if len(vals) < 3:
        return None

    extrema: List[float] = []
    for i in range(1, len(vals) - 1):
        prev, cur, nxt = vals[i - 1], vals[i], vals[i + 1]
        if cur >= prev and cur >= nxt and not (cur == prev == nxt):
            extrema.append(cur)
        elif cur <= prev and cur <= nxt and not (cur == prev == nxt):
            extrema.append(cur)

    if len(extrema) < 2:
        return None

    diffs = [abs(extrema[i + 1] - extrema[i]) for i in range(len(extrema) - 1)]
    if not diffs:
        return None
    return statistics.mean(diffs)


def format_value(value: Optional[float]) -> str:
    return "n/a" if value is None else f"{value:.6f}"


def plot_deltas(times: List[float], deltas: List[float], title: str, output_path: Path, yzoom: Optional[float] = None) -> None:
    if plt is None:
        print("matplotlib is not installed; cannot create plot.")
        return

    x_values = times[1:]
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(x_values, deltas, marker="o", linestyle="-", color="tab:blue")
    ax.axhline(0.0, color="tab:red", linestyle="--", linewidth=0.8)
    ax.set_xlabel("TIME (sec)")
    ax.set_ylabel("Delta TIME (sec)")
    ax.set_title(title)

    mean_delta, deviation = compute_statistics(deltas)
    if mean_delta is None:
        mean_delta = 0.0
    if deviation is None:
        deviation = 0.0

    if yzoom is not None:
        span = float(yzoom)
    else:
        span = max(5 * (deviation or 0.0), 0.002)

    ax.set_ylim(mean_delta - span, mean_delta + span)
    try:
        import matplotlib.ticker as mticker
        ax.yaxis.set_major_formatter(mticker.FormatStrFormatter("%.6f"))
        ax.yaxis.set_minor_locator(mticker.AutoMinorLocator())
    except Exception:
        pass

    ax.grid(True, linestyle="--", alpha=0.5)
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)
    print(f"  Plot saved to {output_path}")


def normalize_cap_channel(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if text.startswith("CAP") and text[3:].isdigit():
        return text
    if text.isdigit():
        channel_num = int(text)
        if 1 <= channel_num <= 8:
            return f"CAP{channel_num}"
    return None


def plot_cap_channels(
    times: List[float],
    cap_data: dict[str, list[Optional[float]]],
    title: str,
    output_path: Path,
    yzoom: Optional[float] = None,
    selected_channel: Optional[str] = None,
) -> None:
    if plt is None:
        print("matplotlib is not installed; cannot create plot.")
        return

    fig, ax = plt.subplots(figsize=(12, 6))
    has_data = False
    channels_to_plot = (
        [selected_channel] if selected_channel is not None else sorted(cap_data.keys(), key=lambda name: int(name[3:]))
    )

    for channel in channels_to_plot:
        if channel not in cap_data:
            continue
        values = cap_data[channel]
        if not any(v is not None for v in values):
            continue
        has_data = True
        ax.plot(times, values, label=channel, linewidth=1.5)

    if not has_data:
        print("  No CAP channel values found; skipping CAP plot.")
        plt.close(fig)
        return

    ax.set_xlabel("TIME (sec)")
    ax.set_ylabel("Capacitance (pF)")
    ax.set_title(title)
    ax.legend(loc="upper right", ncol=2, fontsize="small")

    if yzoom is not None:
        all_values = [v for values in cap_data.values() for v in values if v is not None]
        if all_values:
            mean_value = statistics.mean(all_values)
            span = float(yzoom)
            ax.set_ylim(mean_value - span, mean_value + span)

    ax.grid(True, linestyle="--", alpha=0.5)
    fig.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)
    print(f"  Plot saved to {output_path}")


def analyze_file(
    file_path: Path,
    time_limit: Optional[float],
    debug: bool,
    plot: bool,
    plot_cap: bool,
    plot_dir: Optional[Path],
    yzoom: Optional[float],
    cap_channel: Optional[str] = None,
) -> None:
    print(f"\nAnalyzing {file_path.name}")
    times, cap_data = parse_sheet_data_from_xlsx(file_path, time_limit=time_limit)
    if not times:
        print("  No TIME values found.")
        return

    deltas = compute_deltas(times)
    mean_delta, deviation = compute_statistics(deltas)

    # Build summary dictionary (will be saved as JSON)
    negative_pairs = find_negative_deltas(times)

    summary: dict = {
        "file": file_path.name,
        "parsed_time_entries": len(times),
        "first_time": times[0],
        "last_time": times[-1],
        "mean_delta": None if mean_delta is None else mean_delta,
        "delta_deviation": None if deviation is None else deviation,
    }

    if deltas:
        summary["min_delta"] = min(deltas)
        summary["max_delta"] = max(deltas)

    # Compute per-channel and mean peak-trough delta across CAP channels (pF)
    cap_deltas: dict[str, float] = {}
    cap_mean_peak_troughs: dict[str, Optional[float]] = {}
    for channel in sorted(cap_data.keys(), key=lambda name: int(name[3:])):
        values = [v for v in cap_data[channel] if v is not None]
        if not values:
            continue
        cap_delta = max(values) - min(values)
        cap_deltas[channel] = cap_delta
        # compute mean of individual peak-trough amplitudes for this channel
        mean_pt = compute_mean_peak_trough(cap_data[channel])
        cap_mean_peak_troughs[channel] = mean_pt

    if cap_deltas:
        summary["cap_peak_trough"] = cap_deltas
        summary["mean_cap_peak_trough"] = statistics.mean(cap_deltas.values())
    else:
        summary["cap_peak_trough"] = {}
        summary["mean_cap_peak_trough"] = None
    # include per-channel mean peak-trough amplitudes (average of per-cycle peak-troughs)
    summary["cap_mean_peak_troughs"] = cap_mean_peak_troughs

    # Include negative deltas in the JSON when debug mode is on
    if debug:
        summary["negative_deltas"] = [
            {"previous": p, "current": c, "delta": d} for p, c, d in negative_pairs
        ]

    # Determine output JSON path
    if plot_dir is None:
        summary_path = file_path.with_suffix("").with_name(f"{file_path.stem}-summary.json")
    else:
        plot_dir.mkdir(parents=True, exist_ok=True)
        summary_path = plot_dir / f"{file_path.stem}-summary.json"

    try:
        with summary_path.open("w", encoding="utf-8") as fh:
            json.dump(summary, fh, indent=2)
        print(f"  Summary saved to {summary_path}")
    except Exception as exc:
        print(f"  Failed to save summary to {summary_path}: {exc}")

    if debug:
        negative_pairs = find_negative_deltas(times)
        if not negative_pairs:
            print("  No negative deltas found.")
        else:
            print("  Negative deltas:")
            for prev_time, current_time, delta in negative_pairs:
                print(f"    previous={format_value(prev_time)} sec, current={format_value(current_time)} sec, delta={format_value(delta)} sec")

    if plot:
        if plot_dir is None:
            output_path = file_path.with_suffix("").with_name(f"{file_path.stem}-delta.png")
        else:
            plot_dir.mkdir(parents=True, exist_ok=True)
            output_path = plot_dir / f"{file_path.stem}-delta.png"
        plot_deltas(times, deltas, f"Delta over time for {file_path.name}", output_path, yzoom=yzoom)

    if plot_cap or cap_channel is not None:
        selected_channel = normalize_cap_channel(cap_channel)
        if cap_channel is not None and selected_channel is None:
            print(f"  Invalid CAP channel requested: {cap_channel}. Use 1-8 or CAP1-CAP8.")
            return

        if plot_dir is None:
            suffix = "-cap-channels" if selected_channel is None else f"-{selected_channel.lower()}"
            output_path = file_path.with_suffix("").with_name(f"{file_path.stem}{suffix}.png")
        else:
            plot_dir.mkdir(parents=True, exist_ok=True)
            suffix = "-cap-channels" if selected_channel is None else f"-{selected_channel.lower()}"
            output_path = plot_dir / f"{file_path.stem}{suffix}.png"

        title = (
            f"{selected_channel} over time for {file_path.name}"
            if selected_channel is not None
            else f"CAP1-8 over time for {file_path.name}"
        )
        plot_cap_channels(times, cap_data, title, output_path, yzoom=yzoom, selected_channel=selected_channel)


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse Excel files and analyze the TIME column.")
    parser.add_argument("folder", type=Path, nargs="?", default=Path("wavegen/"), help="Folder containing .xlsx files to analyze.")
    parser.add_argument("--limit", type=float, default=None, help="Optional TIME limit in seconds.")
    parser.add_argument("--debug", action="store_true", help="Print negative TIME deltas for each file.")
    parser.add_argument("--plot", action="store_true", help="Save plots of delta over time for each file.")
    parser.add_argument("--plot-cap", action="store_true", help="Save plots of CAP1..CAP8 over TIME for each file.")
    parser.add_argument("--cap-channel", type=str, default=None, help="Optional single CAP channel to plot, e.g. 1 or CAP3.")
    parser.add_argument("--yzoom", type=float, default=None, help="Half-range for y-axis in seconds (e.g. 0.001). Overrides automatic scaling.")
    parser.add_argument("--plot-dir", type=Path, default=None, help="Optional directory to save plot files.")
    args = parser.parse_args()

    folder = args.folder
    if not folder.exists() or not folder.is_dir():
        raise SystemExit(f"Folder not found: {folder}")

    files = list(get_xlsx_files(folder))
    if not files:
        raise SystemExit(f"No .xlsx files found in {folder}")

    for file_path in files:
        analyze_file(
            file_path,
            time_limit=args.limit,
            debug=args.debug,
            plot=args.plot,
            plot_cap=args.plot_cap,
            plot_dir=args.plot_dir,
            yzoom=args.yzoom,
            cap_channel=args.cap_channel,
        )


if __name__ == "__main__":
    main()
